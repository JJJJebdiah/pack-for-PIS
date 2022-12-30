'''
python 3.10.0
'''

import openpyxl
import os


class Excel:
    def __init__(self):
        self.filename = ''  # 文件名
        self.type = ''  # 出/入库类型
        self.date = ''  # 日期
        self.unit = ''  # 客户名称
        self.rds = []
        pass

    pass


def readfile():
    e = Excel()
    with open(os.getcwd() + r'\DialogImport.dat', 'r', encoding='utf-8') as f:
        # 逐行读取
        e.type = f.readline()[0:-1]
        e.filename = f.readline()[0:-1]
        e.date = f.readline()[0:-1]
        e.unit = f.readline()[0:-1]
        line = f.readline()[0:-1]
        while line:
            rd = line.split(' ')
            rd.insert(0, str(len(e.rds) + 1))
            e.rds.append(rd)
            line = f.readline()
            pass
        pass
    return e
    pass


def create_excel(e):
    workbook = openpyxl.Workbook()  # 创建一个Workbook对象，相当于创建了一个Excel文件
    worksheet = workbook.active  # 获取当前活跃的worksheet,默认就是第一个worksheet
    worksheet.title = 'Sheet1'  # 给当前活跃的worksheet命名
    worksheet.merge_cells('A1:E1')  # 将第一行的前五列合并
    if e.type == 'import':
        worksheet['A1'] = '徐州捷发机电科技有限公司入库单'
        pass
    else:
        worksheet['A1'] = '徐州捷发机电科技有限公司发货单'
        pass
    worksheet['A1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')  # 设置单元格居中
    worksheet['A1'].font = openpyxl.styles.Font(name='宋体', size=20, bold=True)  # 设置字体
    worksheet.merge_cells('A2:B2')
    worksheet['A2'] = '日期：' + e.date
    worksheet['A2'].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
    worksheet['A2'].font = openpyxl.styles.Font(name='宋体', size=12, bold=True)
    worksheet.merge_cells('D2:E2')
    worksheet['D2'] = '客户名称：' + e.unit
    worksheet['D2'].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
    worksheet['D2'].font = openpyxl.styles.Font(name='宋体', size=12, bold=True)
    if e.type == 'import':
        worksheet.append(['序号', '物料编码', '图号名称', '入库数量', '备注'])  # 写入一行
        pass
    else:
        worksheet.append(['序号', '物料编码', '图号名称', '出库数量', '备注'])
        pass
    width = [10, 18, 18, 12, 20]
    for i in e.rds:
        if width[1] < len(i[3]) + 1:
            width[1] = len(i[3]) + 1
            pass
        if width[2] < len(i[1]) + 1:
            width[2] = len(i[1]) + 1
            pass
        if width[3] < len(i[2]) + 1:
            width[3] = len(i[2]) + 1
            pass
        if width[4] < (len(i[4]) + 1) * 2:
            width[4] = (len(i[4]) + 1) * 2
            pass
        worksheet.append(i)
        pass
    # 设置单元格宽度
    for i in range(len(width)):
        worksheet.column_dimensions[chr(ord('A') + i)].width = width[i]
        pass
    #获得文件行数
    row = worksheet.max_row
    worksheet['A' + str(row + 2)] = '发货人：'
    worksheet['C' + str(row + 2)] = '装车人：'
    worksheet['E' + str(row + 2)] = '收货人：'
    workbook.save(e.filename)
    pass


if __name__ == '__main__':
    e = readfile()
    create_excel(e)
    pass
